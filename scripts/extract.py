#!/usr/bin/env python3
"""
Extract Trust_Structures_Lookup_v2_3.xlsx into structured JSON under src/data/.

The workbook is the single source of truth for all trust content. Re-run this
after editing the spreadsheet:  python3 scripts/extract.py

Outputs (src/data/):
  trusts.json        - 9 core lifetime-transfer trusts (Exec Summary matrix)
  situational.json   - 7 situational / at-death trusts (Situational Trusts)
  situations.json    - 19 client situations (Client Situations)
  objectives.json    - 20 decision-tree objectives (Decision Tree)
  glossary.json      - 40 glossary terms (Glossary)
  figures.json       - key federal figures + state panel (Exec Summary tail)
  explainers.json    - GST + ETIP long-form explainers
"""
import json
import re
from pathlib import Path
import openpyxl

HERE = Path(__file__).resolve().parent
REPO = HERE.parent
SRC = REPO / "src" / "data"
SRC.mkdir(parents=True, exist_ok=True)
XLSX = HERE / "Trust_Structures_Lookup_v2_3.xlsx"

wb = openpyxl.load_workbook(XLSX, data_only=True)


def cell(ws, r, c):
    v = ws.cell(row=r, column=c).value
    if v is None:
        return ""
    return str(v).strip()


# ---------------------------------------------------------------------------
# Core nine trusts  (Exec Summary rows 1-26, columns 2-10)
# The sheet is attribute-per-row, trust-per-column. We transpose it.
# ---------------------------------------------------------------------------
# Curated identity for each core trust column (col index -> id/abbr).
# SLANT sits in column 3, immediately after the SLAT it is engineered from.
CORE_IDS = {
    2: ("slat", "SLAT"),
    3: ("slant", "SLANT"),
    4: ("ilit", "ILIT"),
    5: ("grat", "GRAT"),
    6: ("idgt", "IDGT"),
    7: ("dynasty", "Dynasty"),
    8: ("crt", "CRT"),
    9: ("qprt", "QPRT"),
    10: ("clt", "CLAT"),
}

# The 26 attribute rows, in sheet order, with a machine key + display label.
# Keys are stable identifiers used by the React components.
CORE_ATTRS = [
    (1, "primaryPurpose", "Primary Purpose"),
    (2, "name", "Trust Name"),
    (3, "howItWorks", "How It Works (with example)"),
    (4, "typicalClient", "Typical Client"),
    (5, "usesExemption", "Uses Lifetime Gift Exemption"),
    (6, "inEstate", "Included in Grantor Estate"),
    (7, "basisStepUp", "Receives Basis Step-up"),
    (8, "incomeTaxPayer", "Income Tax Payer"),
    (9, "failureModes", "Top Failure Modes & Drafting Traps"),
    (10, "mortalityRisk", "Grantor Mortality Risk to Trust"),
    (11, "bestAssets", "Best Assets"),
    (12, "gstStatus", "GST Status & Allocation Timing"),
    (13, "qsbsFit", "QSBS Stacking Fit"),
    (14, "hurdleRate", "Hurdle / Valuation Rate Used"),
    (15, "flexibility", "Ongoing Flexibility — Add / Sell / Borrow Later"),
    (16, "assetProtection", "Asset Protection"),
    (17, "swapPower", "Swap Power"),
    (18, "annualCost", "Typical Annual Maintenance Cost (2026)"),
    (19, "setupCost", "Typical Legal Setup Cost (2026)"),
    (20, "ageToImplement", "Typical Age to Implement"),
    (21, "duration", "Typical Duration"),
    (22, "beneficiaries", "Primary Beneficiaries"),
    (23, "fundingMethod", "Funding Method"),
    (24, "complexity", "Drafting Complexity"),
    (25, "popularity", "Popularity"),
    (26, "combinations", "Typical Combination with Other Trusts & Strategies"),
]


def extract_core():
    ws = wb["Exec Summary"]
    trusts = []
    for col, (tid, abbr) in CORE_IDS.items():
        attrs = {}
        for row, key, label in CORE_ATTRS:
            attrs[key] = cell(ws, row, col)
        trusts.append(
            {
                "id": tid,
                "abbr": abbr,
                "name": attrs["name"],
                "primaryPurpose": attrs["primaryPurpose"],
                "attrs": attrs,
            }
        )
    # The ordered attribute schema travels with the data so the UI can group it.
    schema = [{"key": k, "label": l} for (_r, k, l) in CORE_ATTRS]
    return {"schema": schema, "trusts": trusts}


# ---------------------------------------------------------------------------
# Situational trusts  (Situational Trusts rows 2-9, columns 2-8)
# ---------------------------------------------------------------------------
SITU_IDS = {
    2: ("credit-shelter", "Credit Shelter"),
    3: ("qtip", "QTIP"),
    4: ("qdot", "QDOT"),
    5: ("snt", "SNT"),
    6: ("dapt", "DAPT"),
    7: ("bdit", "BDIT"),
    8: ("ing", "ING"),
}

SITU_ATTRS = [
    (2, "name", "Trust"),
    (3, "whatItIs", "What It Is"),
    (4, "irrevocableWhen", "Irrevocable? (and when it becomes so)"),
    (5, "trigger", "Trigger — When It Enters the Conversation"),
    (6, "keyMechanics", "Key Mechanics"),
    (7, "transferTaxRole", "Transfer-Tax Role"),
    (8, "watchOuts", "Watch-outs & Failure Modes"),
    (9, "notInCoreNine", "Why It Is Not in the Core Nine"),
]


def extract_situational():
    ws = wb["Situational Trusts"]
    # Row 1 holds two category banners spanning the columns.
    cat_death = cell(ws, 1, 2)
    cat_special = cell(ws, 1, 5)
    categories = {"death": cat_death, "special": cat_special}
    trusts = []
    for col, (tid, abbr) in SITU_IDS.items():
        attrs = {}
        for row, key, label in SITU_ATTRS:
            attrs[key] = cell(ws, row, col)
        category = "death" if col <= 4 else "special"
        trusts.append(
            {
                "id": tid,
                "abbr": abbr,
                "name": attrs["name"],
                "category": category,
                "attrs": attrs,
            }
        )
    schema = [{"key": k, "label": l} for (_r, k, l) in SITU_ATTRS]
    return {"schema": schema, "categories": categories, "trusts": trusts}


# ---------------------------------------------------------------------------
# Client situations  (Client Situations rows 2-19)
# ---------------------------------------------------------------------------
def slug(text):
    s = re.sub(r"[^a-z0-9]+", "-", text.lower()).strip("-")
    return s[:60]


def extract_situations():
    ws = wb["Client Situations"]
    out = []
    for r in range(2, ws.max_row + 1):
        situation = cell(ws, r, 1)
        if not situation:
            continue
        out.append(
            {
                "id": slug(situation),
                "situation": situation,
                "trusts": cell(ws, r, 2),
                "whyItFits": cell(ws, r, 3),
                "watchOuts": cell(ws, r, 4),
            }
        )
    return out


# ---------------------------------------------------------------------------
# Decision-tree objectives  (Decision Tree rows 2-20)
# ---------------------------------------------------------------------------
def extract_objectives():
    ws = wb["Decision Tree"]
    out = []
    for r in range(2, ws.max_row + 1):
        obj = cell(ws, r, 1)
        if not obj:
            continue
        out.append(
            {
                "id": slug(obj),
                "objective": obj,
                "startWith": cell(ws, r, 2),
                "reason": cell(ws, r, 3),
            }
        )
    return out


# ---------------------------------------------------------------------------
# Glossary  (rows 2-36)
# ---------------------------------------------------------------------------
def extract_glossary():
    ws = wb["Glossary"]
    out = []
    for r in range(2, ws.max_row + 1):
        term = cell(ws, r, 1)
        if not term:
            continue
        out.append({"term": term, "definition": cell(ws, r, 2)})
    return out


# ---------------------------------------------------------------------------
# Key federal figures + state panel  (Exec Summary rows 28-42)
# ---------------------------------------------------------------------------
def extract_figures():
    ws = wb["Exec Summary"]
    figures = []
    # rows 29-35: label(col1) / value(col3) / note(col4)
    for r in range(29, 36):
        label = cell(ws, r, 1)
        if not label:
            continue
        figures.append(
            {"label": label, "value": cell(ws, r, 3), "note": cell(ws, r, 4)}
        )
    figures_intro = cell(ws, 28, 1)
    # rows 38-40: state panel (label col1 / body col2)
    states = []
    for r in range(38, 41):
        name = cell(ws, r, 1)
        if not name:
            continue
        states.append({"state": name, "body": cell(ws, r, 2)})
    sources = cell(ws, 42, 1)
    return {
        "intro": figures_intro,
        "figures": figures,
        "states": states,
        "sources": sources,
    }


# ---------------------------------------------------------------------------
# GST + ETIP explainers  (label col1 / body col2, section per row)
# ---------------------------------------------------------------------------
def extract_explainer(sheet):
    ws = wb[sheet]
    title = cell(ws, 1, 1)
    subtitle = cell(ws, 2, 1)
    sections = []
    for r in range(3, ws.max_row + 1):
        heading = cell(ws, r, 1)
        body = cell(ws, r, 2)
        if not heading and not body:
            continue
        sections.append({"heading": heading, "body": body})
    return {"title": title, "subtitle": subtitle, "sections": sections}


def extract_explainers():
    return {
        "gst": extract_explainer("GST Explainer"),
        "etip": extract_explainer("ETIP Explainer"),
    }


def write(name, data):
    path = SRC / name
    path.write_text(json.dumps(data, indent=2, ensure_ascii=False) + "\n")
    print(f"  wrote {name}")


def main():
    print("Extracting workbook ->", SRC)
    write("trusts.json", extract_core())
    write("situational.json", extract_situational())
    write("situations.json", extract_situations())
    write("objectives.json", extract_objectives())
    write("glossary.json", extract_glossary())
    write("figures.json", extract_figures())
    write("explainers.json", extract_explainers())
    print("Done.")


if __name__ == "__main__":
    main()
